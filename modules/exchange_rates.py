import requests
import openpyxl
import certifi

# API URL
api_url = "https://v6.exchangerate-api.com/v6/4fd120e4fbf752644f2d3019/latest/USD"

# Spreadsheet
spreadsheet_dir = "exchange_rates.xlsx"

#Fetches exchange rates from API and updates & saves spreadsheet
def update_rates():
    response = requests.get(api_url, verify=certifi.where())
    response.raise_for_status()

    data = response.json()
    rates = data.get("conversion_rates", {})

    try:
        wb = openpyxl.load_workbook(spreadsheet_dir)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active

    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

    for i, (currency, rate) in enumerate(rates.items(), start=2):
        sheet.cell(row=i, column=1, value=currency)
        sheet.cell(row=i, column=2, value=rate)

    wb.save(spreadsheet_dir)

#creates a dictionary of all the units from the spreadsheet
def create_dict():
    workbook = openpyxl.load_workbook(filename=spreadsheet_dir)
    sheet = workbook.active
    
    data_dict = {}
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
       data_dict[row[0]] = row[1]
    
    return data_dict

# if __name__ == "__main__":
