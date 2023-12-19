import requests
import openpyxl
import certifi

# API URL
API_URL = "https://v6.exchangerate-api.com/v6/4fd120e4fbf752644f2d3019/latest/USD"

# Spreadsheet
SPREADSHEET_FILE = "modules/exchange_rates.xlsx"

#Fetches exchange rates from API and updates spreadsheet
def update_rates():
    # Send API request
    response = requests.get(API_URL, verify=certifi.where())
    response.raise_for_status()

    # Get rates from response
    data = response.json()
    rates = data.get("conversion_rates", {})

    # Create a new spreadsheet or load an existing one
    try:
        wb = openpyxl.load_workbook(SPREADSHEET_FILE)
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active

    # Prune the spreadsheet
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

    # Write rates to the spreadsheet
    for i, (currency, rate) in enumerate(rates.items(), start=2):
        sheet.cell(row=i, column=1, value=currency)
        sheet.cell(row=i, column=2, value=rate)

    # Save changes on spreadsheet
    wb.save(SPREADSHEET_FILE)

#creates a dictionary of all the units from the spreadsheet
def create_dict():
    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(filename=SPREADSHEET_FILE)
    sheet = workbook.active
    
    # Initialize an empty dictionary
    data_dict = {}
    
    # Iterate over the rows in the sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
       data_dict[row[0]] = row[1]

    # Print the dictionary
    print(data_dict)
    
if __name__ == "__main__":
    update_rates()
    create_dict()
